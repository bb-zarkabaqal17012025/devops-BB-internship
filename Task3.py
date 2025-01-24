import re
import boto3
from openpyxl import load_workbook

input_workbook = load_workbook("missing_tags_instances.xlsx")
input_sheet = input_workbook.active

key_name_pattern = re.compile(r"^(?!aws:)[\w\s\.\:\+\=\@\_\/-]{1,128}$")
instance_id_pattern = re.compile(r"^i-[a-f0-9]{8,17}$")

input_data = {}
for row in input_sheet.iter_rows(min_row=2, values_only=True):
    instance_id = row[0]
    if instance_id and instance_id_pattern.match(instance_id):
        input_data[instance_id] = {
            'InstanceType': row[1] if row[1] is not None else "",
            'Name': row[2] if row[2] is not None else "",
            'application_module': row[3] if row[3] is not None else "",
            'team': row[4] if row[4] is not None else "",
            'patch': row[5] if row[5] is not None else "",
            'os': row[6] if row[6] is not None else "",
            'ssm': row[7] if row[7] is not None else "",
            'State': row[8] if row[8] is not None else "",
        }
    elif instance_id:
        print(f"Invalid instance ID format: {instance_id}")

regions = {"ap-south-1", "us-east-1"}

for region in regions:
    ec2 = boto3.client('ec2', region_name=region)
    response = ec2.describe_instances()
    
    for res in response['Reservations']:
        for instance in res['Instances']:
            instance_id = instance['InstanceId']
            instance_state = instance['State']['Name']
            
            if instance_id in input_data and instance_state == 'running':
                instance_tags_dict = {tag['Key']: tag['Value'] for tag in instance.get('Tags', [])}
                required_tags = input_data[instance_id]
                valid_tag_keys = {'Name', 'application_module', 'team', 'patch', 'os', 'ssm'}
                
                missing_or_empty_tags = {
                    key: value for key, value in required_tags.items()
                    if (
                        key in valid_tag_keys and
                        key_name_pattern.match(key) and
                        (key not in instance_tags_dict or instance_tags_dict[key] == "") and
                        value != "" and
                        len(value.encode('utf-8')) <= 256
                    )
                }
                
                invalid_values = {
                    key: value for key, value in required_tags.items()
                    if key in valid_tag_keys and len(value.encode('utf-8')) > 256
                }
                if invalid_values:
                    print(f"Error: Tag values exceed 256 characters for instance {instance_id}: {invalid_values}")
                
                unique_keys = set()
                final_tags = []
                for key, value in missing_or_empty_tags.items():
                    if key not in unique_keys:
                        unique_keys.add(key)
                        final_tags.append({'Key': key, 'Value': value})
                
                if final_tags:
                    try:
                        ec2.create_tags(
                            Resources=[instance_id],
                            Tags=final_tags
                        )
                        print(f"Updated tags for instance {instance_id}: {final_tags}")
                    except Exception as e:
                        print(f"Error updating tags for instance {instance_id}: {e}")
