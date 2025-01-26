import boto3
from openpyxl import Workbook
regions = ["ap-south-1", "us-east-1"]

workbook = Workbook()
def sort_by_start_time(snapshot):
    return snapshot.get('StartTime', '9999-12-31') #if start time  isn not available

for region in regions:
    ec2 = boto3.client('ec2', region_name=region)
    all_snapshots = ec2.describe_snapshots(OwnerIds=['self'])['Snapshots']
    ami_snapshots = set()  
    images = ec2.describe_images(Owners=['self'])['Images']
    for image in images:
        for block in image.get('BlockDeviceMappings', []):
            if 'Ebs' in block and 'SnapshotId' in block['Ebs']:
                ami_snapshots.add(block['Ebs']['SnapshotId'])
    unlinked_snapshots = [
        {
            'SnapshotId': snapshot['SnapshotId'],
             'VolumeId': snapshot.get('VolumeId', 'N/A'),
            'Size': snapshot.get('VolumeSize', 'N/A'),
            'StartTime': snapshot['StartTime'].strftime('%Y-%m-%d %H:%M:%S')

        }
        for snapshot in all_snapshots
        if snapshot['SnapshotId'] not in ami_snapshots
    ]
    unlinked_snapshots = sorted(unlinked_snapshots, key=sort_by_start_time)
    sheet = workbook.create_sheet(title=region)
    headers = ['Snapshot Id', 'Volume Id', 'Size (GiB)', 'Start time']
    sheet.append(headers)
    for snapshot in unlinked_snapshots:
        sheet.append([
            snapshot['SnapshotId'],
            snapshot['VolumeId'],
            snapshot['Size'],
            snapshot['StartTime']
        ])
if "Sheet" in workbook.sheetnames:
    del workbook["Sheet"]
output_file = "unlinked_snapshots_by_region.xlsx"
workbook.save(output_file)
print(f"Unlinked snapshots saved to {output_file}")
